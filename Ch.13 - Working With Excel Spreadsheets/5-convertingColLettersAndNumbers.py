# Converting between column letters and numbers

import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string

# Turn num into col letter
print(get_column_letter(1))
print(get_column_letter(2))

wb = xl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

print(get_column_letter(sheet.max_column))

# Get num from col letter
print(column_index_from_string('A'))
print(column_index_from_string('AA'))
