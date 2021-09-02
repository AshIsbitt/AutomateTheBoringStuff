# Looping through cells 

import openpyxl as xl

wb = xl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

# Cells 1-8, every other cell
for i in range(1, 8, 2):
	print(i, sheet.cell(row=i, column=2).value)
