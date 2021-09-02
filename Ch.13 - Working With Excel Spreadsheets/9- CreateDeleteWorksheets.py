# Creating and removing sheets

import openpyxl as xl

wb = xl.Workbook('example_2.xlsx')
print(wb.sheetnames)

# Add sheet
# This creates a new Workbook Object
wb.create_sheet()
print(wb.sheetnames)

# Create a sheet with a specific name and location
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)

# Delete sheet
del wb['Sheet']
print(wb.sheetnames)

wb.save('example_3.xlsx')
