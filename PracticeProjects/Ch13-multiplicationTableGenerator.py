# multiplication table generator

'''
Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication 
table in an Excel spreadsheet. For example, when the program is run like this:
py multiplicationTable.py 6
. . . it should create a spreadsheet that looks like Figure 13-11
'''

import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sys

# Get input in command line
multiplier = 0

if len(sys.argv) == 2:
	multiplier = int(sys.argv[1])
else:
	sys.exit()

# Create new worksheet
wb = xl.Workbook()
sheet = wb.active

# Set fonts for first row/col
sidebarFont = Font(bold=True, size=16)

# Create rows and cols
print('Generating Table...')

for i in range(1, multiplier+1):
	sheet[f'A{i+1}'] = i
	sheet[f'A{i+1}'].font = sidebarFont

	sheet[f'{get_column_letter(i+1)}1'] = i
	sheet[f'{get_column_letter(i+1)}1'].font = sidebarFont


# Populating with maths
print('Populating with data...')

# Rows
for i in range(2, multiplier+2):
	# Cols
	for j in range(2, multiplier+2):
		val = sheet[f'A{i}'].value*sheet[f'{get_column_letter(j)}1'].value

		# Put val in spreadsheet
		sheet[f'{get_column_letter(j)}{i}'] = val


# Save table
wb.save('MultiplicationTable.xlsx')