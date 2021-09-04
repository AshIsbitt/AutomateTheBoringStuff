# Excel-to-CSV Converter

import openpyxl
from openpyxl.utils import get_column_letter
import csv
import os
from pathlib import Path

for excelFile in os.listdir('./excelSpreadsheets/'):
    # Skip non-excel files
    p = Path(f'./excelSpreadsheets/{excelFile}')

    if p.suffix != '.xlsx':
        continue

    print(f'Opening {excelFile}...')

    # Set spreadsheet as Spreadsheet object
    wb = openpyxl.load_workbook(os.path.join('excelSpreadsheets', excelFile))

    # Create location for output
    Path('./excelSpreadsheets/outputCSVfiles/').mkdir(exist_ok=True)

    # Loop through all sheets in workbook
    for sheetName in wb.sheetnames:
        # Get sheet
        sheet = wb[sheetName]

        # Create CSV filename from excel filename and sheet name
        filename = f"{str(excelFile[:-5])}-{str(sheet.title)}"
        print(filename)

        # Create CSV writer
        csvFile = open(
            f'./excelSpreadsheets/outputCSVfiles/{filename}.csv', 'w')
        csvWriter = csv.writer(csvFile)

        # Loop through every row in the sheet
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []

            for colNum in range(1, sheet.max_column + 1):
                # Append each cell to rowData

                rowData.append(
                    sheet[f'{get_column_letter(colNum)}{rowNum}'].value)

            # write rowData list to CSV file
            print(f'rowdata = {rowData}')
            csvWriter.writerow(rowData)

    csvFile.close()

print('Finished!')
