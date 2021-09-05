# Project - Sending Members Dues

import openpyxl
import smtplib
import sys
import USERNAMEANDPASSWORD as upw

# Open spreadsheet and get latest dues status
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1. column=lastCol).value

# Check each member's payment status
unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value

    if payment != 'paid':
        name = sheet.cell(row=r, column=1)
        email = sheet.cell(row=r, column=2)
        unpaidMembers[name] = email

# log in to email account
smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(upw.email, upw.password)

# Send out reminder emails
for name, email in unpaidMembers.items():
    body = f"""Subject {latestMonth} dues unpaid. \n\n Dear {name}, Records show that you have not
    paid dues for {latestMonth}. Please make this payment as soon as possible. Thank you!"""

    print(f"Sending email to {email}")
    sendmailStatus = smtpObj.sendmail(upw.email, email, body)

    if sendmailStatus != {}:
        print(f"There was a problem sending to {email}: {sendmailStatus}")

smtpObj.quit()
