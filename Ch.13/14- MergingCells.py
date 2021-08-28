# Merging/unmerging cells

import openpyxl as xl
wb = xl.Workbook()
sheet = wb.active

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged'

sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'

wb.save('merged.xlsx')
